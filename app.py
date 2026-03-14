# ================================================================
# AI SUPPORT REPLY GENERATOR  ·  LangGraph + RAG + Groq
# Multi-Industry · Compact UI · v3.0
# NEW: Auto Priority from past tickets · Email fallback on no KB match
# ================================================================
import os, html, time, datetime, smtplib, base64, json
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path
from typing import TypedDict
import tempfile as _tmp

import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from langchain_groq import ChatGroq
from langchain_core.prompts import ChatPromptTemplate
from langchain_core.output_parsers import StrOutputParser
from langchain_community.document_loaders import PyPDFLoader
from langchain_community.vectorstores import FAISS
from langchain_huggingface import HuggingFaceEmbeddings
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langgraph.graph import StateGraph, END

# ── Config ───────────────────────────────────────────────────
BASE_DIR             = _tmp.gettempdir()
GROQ_API_KEY         = os.environ.get("GROQ_API_KEY", "")
PDF_PATH             = os.environ.get("PDF_PATH", "")
EMBED_MODEL          = "sentence-transformers/all-MiniLM-L6-v2"
TOP_K                = 5
EXCEL_PATH           = os.path.join(BASE_DIR, "support_tickets.xlsx")
HISTORY_PATH         = os.path.join(BASE_DIR, "ticket_history.json")
SMTP_EMAIL           = os.environ.get("SMTP_EMAIL", "")
SMTP_PASS            = os.environ.get("SMTP_PASSWORD", "")

# ── Industry Profiles ────────────────────────────────────────
INDUSTRIES = {
    "🏫 Art of Living":  {"a":"#E8540A","a2":"#F97316","bg":"#090400","s":"#120700","greeting":"JGD,","sign":"AOL Teacher Support","label":"Teacher","code_label":"Teacher Code","org":"artofliving.online"},
    "🏥 Healthcare":     {"a":"#0EA5E9","a2":"#38BDF8","bg":"#00040A","s":"#00091A","greeting":"Dear,","sign":"Patient Support Team","label":"Patient","code_label":"Patient ID","org":"healthsupport.org"},
    "🏦 Banking":        {"a":"#6366F1","a2":"#818CF8","bg":"#02010A","s":"#05041A","greeting":"Dear,","sign":"Customer Support","label":"Customer","code_label":"Account No","org":"banksupport.com"},
    "🛒 E-Commerce":     {"a":"#10B981","a2":"#34D399","bg":"#00080A","s":"#001208","greeting":"Hi,","sign":"Support Team","label":"Customer","code_label":"Order ID","org":"support@store.com"},
    "🎓 Education":      {"a":"#F59E0B","a2":"#FCD34D","bg":"#080300","s":"#100800","greeting":"Hello,","sign":"Academic Support","label":"Student","code_label":"Student ID","org":"eduhelp.ac.in"},
    "🏭 Manufacturing":  {"a":"#94A3B8","a2":"#CBD5E1","bg":"#03040A","s":"#06080F","greeting":"Dear,","sign":"Technical Support","label":"Engineer","code_label":"Employee ID","org":"techsupport.com"},
    "✈️ Travel":         {"a":"#F472B6","a2":"#FB7185","bg":"#0A0005","s":"#180010","greeting":"Hello,","sign":"Travel Support","label":"Traveler","code_label":"Booking ID","org":"travelsupport.com"},
    "🏗️ Real Estate":    {"a":"#FB923C","a2":"#FDBA74","bg":"#0A0400","s":"#180A00","greeting":"Dear,","sign":"Property Support","label":"Client","code_label":"Client ID","org":"realestate.com"},
}

CATEGORIES_BY_INDUSTRY = {
    "🏫 Art of Living":  ["Login / Authentication","Events","Course Management","Resources","Quotes","Account / Profile","Live Darshan","Manage Addresses","Contact Us","App Performance","Network","Other"],
    "🏥 Healthcare":     ["Appointment Booking","Billing / Insurance","Prescription","Lab Results","Telemedicine","Patient Records","Medication","Emergency","Portal Access","Other"],
    "🏦 Banking":        ["Account Access","Fund Transfer","Card Issues","Loan","KYC / Documents","Interest / Charges","Fraud Alert","UPI / Payment","Net Banking","Other"],
    "🛒 E-Commerce":     ["Order Tracking","Return / Refund","Payment Failed","Product Quality","Delivery Delay","Account Access","Coupon / Offer","Cancel Order","Seller Issue","Other"],
    "🎓 Education":      ["Exam Registration","Result / Grading","Fee Payment","Course Access","Library","Hostel","Scholarship","Certificate","Staff Complaint","Other"],
    "🏭 Manufacturing":  ["Machine Downtime","Quality Issue","Safety Incident","Supply Chain","Maintenance","Production Delay","Compliance","Inventory","Vendor Issue","Other"],
    "✈️ Travel":         ["Flight Booking","Hotel Booking","Cancellation","Baggage","Check-in","Visa Support","Refund","Loyalty Points","Upgrade","Other"],
    "🏗️ Real Estate":    ["Property Listing","Site Visit","Documentation","Loan Assist","Maintenance","Tenant Issue","Legal Query","Valuation","Handover","Other"],
}

QUICK_TIPS = {
    "Login / Authentication": ["Cannot login","Account locked","OTP failed","Session expired"],
    "Events":                 ["Create event fails","Wrong date","Status stuck","Cannot edit"],
    "App Performance":        ["App crashing","Very slow","No notifications","Battery drain"],
    "Payment Failed":         ["Card declined","UPI failed","Timeout","Double charge"],
    "Order Tracking":         ["Order missing","Wrong status","Delayed","No update"],
    "Appointment Booking":    ["Slot unavailable","Booking failed","Wrong doctor","Reschedule"],
}

# ── Priority Options (🤖 Auto = LLM assigns based on past tickets) ────────────
PRIORITIES   = ["🔴 Critical","🟠 High","🟡 Medium","🟢 Low","🤖 Automatic"]
PRIORITY_SLA = {"🔴 Critical":"2 hrs","🟠 High":"4 hrs","🟡 Medium":"24 hrs","🟢 Low":"48 hrs","🤖 Automatic":"LLM"}
PCOLORS      = {"🔴 Critical":"#EF4444","🟠 High":"#F97316","🟡 Medium":"#EAB308","🟢 Low":"#22C55E","🤖 Automatic":"#A78BFA"}

llm = ChatGroq(model="llama-3.1-8b-instant", temperature=0, api_key=GROQ_API_KEY)

# ── Helpers ──────────────────────────────────────────────────
def load_history():
    try: return json.loads(Path(HISTORY_PATH).read_text()) if Path(HISTORY_PATH).exists() else []
    except: return []

def save_history(d):
    h=load_history(); h.insert(0,d); Path(HISTORY_PATH).write_text(json.dumps(h[:100],indent=2))

class TicketState(TypedDict):
    ticket:str; user_name:str; user_email:str; user_code:str
    category:str; priority:str; auto_priority:str; context:str
    kb_matched:bool; raw_response:str; final_response:str; ticket_id:str
    email_fallback:bool

def init_excel():
    if Path(EXCEL_PATH).exists(): return
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Tickets"
    heads=["#","Ticket ID","Time","Name","Email","Code","Category","Priority","Auto-P","Issue","Status","Rating"]
    ws.append(heads)
    hf=PatternFill("solid",fgColor="1E293B"); hfont=Font(bold=True,color="FFFFFF",name="Calibri",size=9)
    thin=Border(*[Side(style="thin")]*4)
    for i in range(1,len(heads)+1):
        c=ws.cell(1,i); c.fill=hf; c.font=hfont
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=thin
    for i,w in enumerate([3,13,18,16,26,12,18,12,12,45,11,7],1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w
    ws.row_dimensions[1].height=22; wb.save(EXCEL_PATH)

def log_excel(ticket,name,email,code,cat,pri,apri,status="Resolved",rating=None):
    init_excel()
    wb=openpyxl.load_workbook(EXCEL_PATH); ws=wb.active; rn=ws.max_row+1
    tid=f"TKT-{int(time.time())%10000000:07d}"
    ts=datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.append([rn-1,tid,ts,name,email,code,cat,pri,apri,ticket,status,rating or "—"])
    thin=Border(*[Side(style="thin")]*4)
    af=PatternFill("solid",fgColor="F8FAFC" if rn%2==0 else "FFFFFF")
    for ci in range(1,13):
        c=ws.cell(rn,ci); c.fill=af; c.border=thin
        c.font=Font(name="Calibri",size=8); c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    wb.save(EXCEL_PATH); return tid

@st.cache_resource(show_spinner=False)
def build_vs(pdf_path):
    docs=PyPDFLoader(pdf_path).load()
    chunks=RecursiveCharacterTextSplitter(chunk_size=500,chunk_overlap=50).split_documents(docs)
    emb=HuggingFaceEmbeddings(model_name=EMBED_MODEL,model_kwargs={"device":"cpu"})
    return FAISS.from_documents(chunks,emb)

def send_email(to_email,to_name,issue,reply,tid,greeting,sign,org,img_b64=None,img_name=None,subject_prefix="[Support] Response"):
    if not SMTP_EMAIL or not SMTP_PASS: return False,"No email credentials"
    try:
        msg=MIMEMultipart("mixed")
        msg["Subject"]=f"{subject_prefix} · {tid}"
        msg["From"]=f"AI Support <{SMTP_EMAIL}>"; msg["To"]=to_email
        alt=MIMEMultipart("alternative")
        plain=f"Dear {to_name},\n\n{reply}\n\n— {sign}"
        rh=html.escape(reply).replace("\n","<br>"); ih=html.escape(issue)
        body=f"""<!DOCTYPE html><html><head><meta charset="UTF-8">
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;600;700&display=swap" rel="stylesheet">
<style>*{{margin:0;padding:0;box-sizing:border-box}}body{{background:#f1f5f9;font-family:'DM Sans',sans-serif}}
.w{{max-width:580px;margin:24px auto;padding:0 12px 24px}}.c{{background:#fff;border-radius:16px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,.08)}}
.h{{background:#0f172a;padding:22px 28px}}.hn{{color:#fff;font-size:18px;font-weight:700}}.hs{{color:rgba(255,255,255,.4);font-size:11px;margin-top:2px}}
.b{{padding:22px 28px}}.lbl{{font-size:9px;font-weight:700;text-transform:uppercase;letter-spacing:2px;color:#94a3b8;margin-bottom:5px}}
.ib{{background:#f8fafc;border-left:3px solid #3b82f6;border-radius:0 8px 8px 0;padding:10px 12px;font-size:13px;color:#334155;line-height:1.6;margin-bottom:16px}}
.rb{{background:#f0fdf4;border-left:3px solid #22c55e;border-radius:0 8px 8px 0;padding:12px 14px;font-size:13px;color:#14532d;line-height:1.7}}
.ft{{background:#f8fafc;padding:12px 28px;text-align:center;border-top:1px solid #e2e8f0;font-size:10px;color:#94a3b8}}
</style></head><body><div class="w"><div class="c">
<div class="h"><div class="hn">AI Support System</div><div class="hs">Ticket {tid} · {org}</div></div>
<div class="b"><div style="font-size:15px;font-weight:700;color:#0f172a;margin-bottom:4px;">Dear {html.escape(to_name)},</div>
<div style="font-size:12px;color:#64748b;margin-bottom:14px;">{html.escape(greeting)} Thank you for contacting us.</div>
<div class="lbl">Your Issue</div><div class="ib">{ih}</div>
<div class="lbl">Our Response</div><div class="rb">{rh}</div></div>
<div class="ft">AI Support · {org}</div></div></div></body></html>"""
        alt.attach(MIMEText(plain,"plain")); alt.attach(MIMEText(body,"html")); msg.attach(alt)
        if img_b64 and img_name:
            p=MIMEBase("application","octet-stream"); p.set_payload(base64.b64decode(img_b64))
            encoders.encode_base64(p); p.add_header("Content-Disposition","attachment",filename=img_name); msg.attach(p)
        with smtplib.SMTP_SSL("smtp.gmail.com",465) as s:
            s.login(SMTP_EMAIL,SMTP_PASS); s.sendmail(SMTP_EMAIL,to_email,msg.as_string())
        return True,f"✅ Email sent to {to_email}"
    except smtplib.SMTPAuthenticationError: return False,"❌ Gmail auth failed"
    except Exception as e: return False,f"❌ {str(e)}"

# ── LangGraph Nodes ───────────────────────────────────────────
def node_classify(state):
    IND=st.session_state.get("industry","🏫 Art of Living")
    cats=CATEGORIES_BY_INDUSTRY.get(IND,CATEGORIES_BY_INDUSTRY["🏫 Art of Living"])
    p=ChatPromptTemplate.from_messages([
        ("system",f"Classify into ONE of: {', '.join(cats)}. Return ONLY the category name."),
        ("human","Ticket: {ticket}")])
    r=(p|llm|StrOutputParser()).invoke({"ticket":state["ticket"]}).strip()
    return {"category":r}

def node_priority(state):
    if "🤖" not in state.get("priority",""):
        return {"auto_priority":state["priority"]}
    hist=load_history()
    past_summary=""
    if hist:
        recent=[h for h in hist[:20] if h.get("auto_priority") or h.get("priority")]
        if recent:
            lines=[f"- [{h.get('category','')}] {h.get('issue','')[:80]} → Priority: {h.get('auto_priority') or h.get('priority','')}" for h in recent[:8]]
            past_summary="Past resolved tickets for reference:\n"+"\n".join(lines)
    p=ChatPromptTemplate.from_messages([
        ("system",
         "You are a support triage specialist. Assign ONE priority level based on the ticket and past patterns.\n"
         "Rules:\n"
         "  Critical = complete system failure, data loss, security breach\n"
         "  High     = major feature broken, no workaround\n"
         "  Medium   = partial issue, workaround exists\n"
         "  Low      = minor/cosmetic/informational\n"
         "Use past ticket patterns to stay consistent. Return ONLY one word: Critical, High, Medium, or Low."),
        ("human","Ticket: {ticket}\nCategory: {category}\n\n{past}")])
    r=(p|llm|StrOutputParser()).invoke({
        "ticket":state["ticket"],
        "category":state.get("category",""),
        "past":past_summary
    }).strip()
    pm={"Critical":"🔴 Critical","High":"🟠 High","Medium":"🟡 Medium","Low":"🟢 Low"}
    clean=next((v for v in pm if v.lower() in r.lower()),"Medium")
    return {"auto_priority":pm[clean]}

def make_retrieve(vstore):
    def node_retrieve(state):
        res=vstore.similarity_search_with_relevance_scores(state["ticket"],k=TOP_K)
        if not res: return {"context":"","kb_matched":False}
        _,bs=res[0]
        return {"context":"\n\n---\n\n".join(d.page_content for d,_ in res),"kb_matched":bs>0.3}
    return node_retrieve

def node_generate(state):
    IND_KEY=st.session_state.get("industry","🏫 Art of Living")
    prof=INDUSTRIES[IND_KEY]
    name=state.get("user_name","") or "there"
    ep=state.get("auto_priority") or state.get("priority","🟡 Medium")
    if not state.get("kb_matched",False):
        tid=log_excel(state["ticket"],name,state.get("user_email","N/A"),
                      state.get("user_code","N/A"),state.get("category","Other"),
                      state.get("priority",""),state.get("auto_priority",""),status="Escalated")
        eta=PRIORITY_SLA.get(ep,"24 hrs")
        escalation_msg=(
            f"{prof['greeting']}\n\nThank you for contacting {prof['sign']}.\n\n"
            f"We were unable to find a matching solution in our knowledge base for your issue. "
            f"Your ticket has been escalated to our specialist team who will personally review and respond.\n\n"
            f"📋 Ticket: {tid}\n"
            f"👤 Name: {name}\n"
            f"🔑 Code: {state.get('user_code','N/A')}\n"
            f"🚨 Priority: {ep}\n"
            f"⏰ ETA: within {eta}\n\n"
            f"For urgent assistance: support@{prof['org']}\n\nThank you"
        )
        return {"raw_response": escalation_msg,"ticket_id": tid,"email_fallback": True}
    p=ChatPromptTemplate.from_messages([
        ("system",f"""You are a senior support specialist for {IND_KEY}.
Always start with '{prof["greeting"]}' and end with 'Thank you'.
Address the user as {name}. Be concise, warm, and professional.
Use numbered steps only when needed. Use past resolved tickets as reference."""),
        ("human","Issue:{ticket}\nCategory:{category}\nPriority:{priority}\n\nPast Tickets:\n{context}")])
    r=(p|llm|StrOutputParser()).invoke({
        "ticket":state["ticket"],"category":state.get("category",""),
        "priority":ep,"context":state.get("context","")}).strip()
    return {"raw_response":r,"email_fallback":False}

def node_polish(state):
    IND_KEY=st.session_state.get("industry","🏫 Art of Living")
    prof=INDUSTRIES[IND_KEY]
    p=ChatPromptTemplate.from_messages([
        ("system",f"Polish this support reply. Ensure it starts with '{prof['greeting']}' and ends with 'Thank you'. Return ONLY the reply."),
        ("human","{r}")])
    r=(p|llm|StrOutputParser()).invoke({"r":state["raw_response"]}).strip()
    return {"final_response":r}

def build_graph(vstore):
    g=StateGraph(TicketState)
    g.add_node("classify",node_classify)
    g.add_node("priority",node_priority)
    g.add_node("retrieve",make_retrieve(vstore))
    g.add_node("generate",node_generate)
    g.add_node("polish",node_polish)
    g.set_entry_point("classify")
    g.add_edge("classify","priority")
    g.add_edge("priority","retrieve")
    g.add_edge("retrieve","generate")
    g.add_edge("generate","polish")
    g.add_edge("polish",END)
    return g.compile()

# ── Page Setup ───────────────────────────────────────────────
st.set_page_config(page_title="AI Support",page_icon="🎯",layout="wide",initial_sidebar_state="expanded")

SESS_DEFAULTS=dict(
    name="",email="",code="",cat_idx=0,priority="🟡 Medium",ticket_input="",
    industry="🏫 Art of Living",vstore=None,graph=None,kb_status="⚠️ No KB",
    category=None,auto_priority=None,kb_matched=None,final_response=None,
    ticket_id=None,email_status=None,current_step=0,
    total=0,resolved=0,escalated=0,emails_sent=0,
    scores=[],chat=[],history=load_history(),img_b64=None,img_name=None,
    email_fallback=False,
)
for k,v in SESS_DEFAULTS.items():
    if k not in st.session_state: st.session_state[k]=v

def reset():
    for k in ["category","auto_priority","kb_matched","final_response",
              "ticket_id","email_status","current_step","email_fallback"]:
        st.session_state[k]=None if k not in ("current_step","email_fallback") else (0 if k=="current_step" else False)

@st.cache_resource(show_spinner=False)
def load_kb(path):
    if os.path.exists(path): return build_vs(path),True
    return None,False

if st.session_state.vstore is None and PDF_PATH:
    with st.spinner("Building KB..."):
        _vs,_ok=load_kb(PDF_PATH)
    if _ok:
        st.session_state.vstore=_vs; st.session_state.graph=build_graph(_vs)
        st.session_state.kb_status=f"✅ {os.path.basename(PDF_PATH)}"; init_excel()

IND   = st.session_state.industry
PROF  = INDUSTRIES[IND]
CATS  = CATEGORIES_BY_INDUSTRY.get(IND,[])
A,A2  = PROF["a"],PROF["a2"]

# ── CSS ──────────────────────────────────────────────────────
st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Geist:wght@300;400;500;600;700;800&family=Geist+Mono:wght@400;500&display=swap');
:root{{
  --a:{A};--a2:{A2};--bg:{PROF['bg']};--s:{PROF['s']};
  --c:rgba(255,255,255,.04);--t:#f1f5f9;--m:#64748b;
  --b:rgba(255,255,255,.07);--gl:{A}40;
  --f:'Geist',sans-serif;--mo:'Geist Mono',monospace;
}}
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
html,body,[class*=css]{{font-family:var(--f)!important;background:var(--bg)!important;color:var(--m)!important}}
[data-testid=stAppViewContainer]{{background:var(--bg)!important}}
[data-testid=stAppViewContainer]::before{{
  content:'';position:fixed;inset:0;pointer-events:none;z-index:0;
  background:
    radial-gradient(ellipse 70% 55% at 8% 0%,{A}0d,transparent 58%),
    radial-gradient(ellipse 50% 45% at 92% 15%,{A2}08,transparent 52%);
}}
[data-testid=stAppViewContainer]>*{{position:relative;z-index:1}}
section[data-testid=stSidebar]{{background:color-mix(in srgb,var(--bg) 97%,{A} 3%)!important;border-right:1px solid var(--b)!important}}
::-webkit-scrollbar{{width:2px;height:2px}}
::-webkit-scrollbar-thumb{{background:{A}44;border-radius:2px}}
.stTabs [data-baseweb=tab-list]{{background:var(--s)!important;border-radius:8px!important;padding:2px!important;gap:2px!important;border:1px solid var(--b)!important}}
.stTabs [data-baseweb=tab]{{background:transparent!important;border-radius:6px!important;color:var(--m)!important;font-family:var(--f)!important;font-weight:600!important;font-size:.75rem!important;padding:6px 14px!important;border:none!important;transition:all .18s!important}}
.stTabs [aria-selected=true]{{background:{A}18!important;color:{A}!important;border:1px solid {A}30!important}}
.stTextArea textarea,.stTextInput>div>div>input,.stSelectbox>div>div{{
  background:color-mix(in srgb,var(--s) 85%,transparent)!important;border:1px solid var(--b)!important;
  border-radius:8px!important;color:var(--t)!important;font-family:var(--f)!important;
  font-size:.82rem!important;transition:all .18s!important;
}}
.stTextArea textarea{{padding:10px 12px!important;line-height:1.55!important}}
.stTextInput>div>div>input{{padding:8px 10px!important;height:38px!important}}
.stTextArea textarea:focus,.stTextInput>div>div>input:focus{{
  border-color:{A}!important;box-shadow:0 0 0 2px {A}25!important;outline:none!important;
}}
.stTextArea textarea::placeholder,.stTextInput>div>div>input::placeholder{{color:{A}35!important}}
.stSelectbox>div>div{{height:38px!important;padding:0 10px!important}}
.stButton>button{{
  background:{A}0e!important;border:1px solid {A}28!important;
  border-radius:7px!important;color:{A}!important;font-family:var(--f)!important;
  font-weight:700!important;font-size:.75rem!important;padding:7px 12px!important;
  transition:all .18s!important;letter-spacing:.1px!important;
}}
.stButton>button:hover{{
  border-color:{A}!important;color:var(--t)!important;background:{A}1a!important;
  box-shadow:0 3px 12px {A}30!important;transform:translateY(-1px)!important;
}}
.stButton>button:disabled{{opacity:.3!important;transform:none!important}}
.stDownloadButton>button{{background:rgba(74,222,128,.07)!important;border:1px solid rgba(74,222,128,.18)!important;color:#4ade80!important;border-radius:7px!important;font-family:var(--f)!important;font-weight:700!important;font-size:.74rem!important}}
.run-btn .stButton>button{{
  background:linear-gradient(135deg,{A}22,{A2}18)!important;
  border-color:{A}55!important;color:var(--t)!important;font-size:.82rem!important;
  padding:11px!important;box-shadow:0 4px 18px {A}30!important;
}}
.run-btn .stButton>button:hover{{box-shadow:0 6px 24px {A}45!important}}
.stSuccess>div{{background:rgba(74,222,128,.05)!important;border:1px solid rgba(74,222,128,.15)!important;border-radius:8px!important;color:#86efac!important;font-size:.78rem!important}}
.stError>div{{background:rgba(239,68,68,.05)!important;border:1px solid rgba(239,68,68,.15)!important;border-radius:8px!important;color:#fca5a5!important;font-size:.78rem!important}}
.stWarning>div{{background:rgba(251,191,36,.05)!important;border:1px solid rgba(251,191,36,.15)!important;border-radius:8px!important;color:#fde68a!important;font-size:.78rem!important}}
.stInfo>div{{background:{A}0a!important;border:1px solid {A}20!important;border-radius:8px!important;color:{A}cc!important;font-size:.78rem!important}}
[data-testid=stFileUploader]>div{{background:{A}06!important;border:1px dashed {A}20!important;border-radius:8px!important}}
.g{{background:var(--c);border:1px solid var(--b);border-radius:12px;padding:14px;margin-bottom:8px;position:relative;overflow:hidden}}
.g::before{{content:'';position:absolute;top:0;left:0;right:0;height:1px;background:linear-gradient(90deg,transparent,{A}12,transparent)}}
.lbl{{font-size:.6rem;font-weight:700;text-transform:uppercase;letter-spacing:1.2px;color:var(--m);margin-bottom:4px;display:flex;align-items:center;gap:4px}}
.dot{{width:3px;height:3px;border-radius:50%;background:{A};opacity:.7}}
.chip{{display:inline-flex;align-items:center;gap:4px;padding:2px 8px;border-radius:100px;font-size:.6rem;font-weight:700;border:1px solid}}
.badge{{display:inline-flex;align-items:center;gap:5px;padding:2px 10px;border-radius:100px;font-size:.6rem;font-weight:800;text-transform:uppercase;letter-spacing:1.5px}}
.hero{{
  border:1px solid {A}14;border-radius:14px;padding:16px 20px;margin-bottom:12px;
  background:linear-gradient(135deg,{PROF['s']}f0,{PROF['bg']}f8);
  position:relative;overflow:hidden;
}}
.hero::before{{content:'';position:absolute;top:0;left:0;right:0;height:1px;background:linear-gradient(90deg,transparent,{A},{A2},transparent);opacity:.3}}
.hero-badge{{color:{A};font-size:.58rem;font-weight:800;text-transform:uppercase;letter-spacing:2px;margin-bottom:8px;display:flex;align-items:center;gap:5px}}
.hero-dot{{width:4px;height:4px;background:{A};border-radius:50%;animation:pulse 2s infinite}}
@keyframes pulse{{0%,100%{{opacity:1;box-shadow:0 0 4px {A}}}50%{{opacity:.25;box-shadow:none}}}}
.hero-h{{font-size:1.45rem;font-weight:700;color:var(--t);letter-spacing:-.3px;line-height:1.2;margin-bottom:4px}}
.hero-h em{{color:{A};font-style:normal}}
.hero-sub{{font-size:.75rem;color:var(--m);line-height:1.5}}
.kpi{{display:flex;gap:0;margin-top:12px;background:rgba(255,255,255,.02);border:1px solid var(--b);border-radius:8px;overflow:hidden}}
.kp{{flex:1;padding:8px 12px;border-right:1px solid var(--b)}}
.kp:last-child{{border-right:none}}
.kv{{font-size:1.1rem;font-weight:700;font-family:var(--mo);line-height:1}}
.kl{{font-size:.52rem;color:var(--m);text-transform:uppercase;letter-spacing:.7px;font-weight:700;margin-top:2px}}
.pipe{{display:flex;align-items:center;gap:2px;flex-wrap:wrap;background:{PROF['s']}cc;border:1px solid var(--b);border-radius:8px;padding:8px 12px;margin:10px 0}}
.pn{{display:inline-flex;align-items:center;gap:4px;padding:4px 8px;border-radius:6px;font-size:.65rem;font-weight:700;border:1.5px solid;white-space:nowrap}}
.pi{{width:14px;height:14px;border-radius:3px;display:flex;align-items:center;justify-content:center;font-size:.52rem;font-weight:900}}
.pn-e,.pn-s{{background:rgba(255,255,255,.02);border-color:rgba(255,255,255,.06);color:var(--m)}}
.pn-d{{background:{A}0e;border-color:{A}25;color:{A}cc}}
.pn-d .pi{{background:{A}18;color:{A}}}
.pn-a{{background:{A}12;border-color:{A};color:{A}f0;box-shadow:0 0 10px {A}35;animation:glow 2s infinite}}
.pn-a .pi{{background:{A};color:{PROF['bg']}}}
.pn-i{{background:rgba(255,255,255,.015);border-color:rgba(255,255,255,.03);color:rgba(255,255,255,.18)}}
.arr{{color:rgba(255,255,255,.1);font-size:.65rem;flex-shrink:0}}
@keyframes glow{{0%,100%{{box-shadow:0 0 6px {A}30}}50%{{box-shadow:0 0 16px {A}55}}}}
.rc{{border-radius:8px;padding:11px 13px;margin:6px 0;position:relative;overflow:hidden}}
.rc::before{{content:'';position:absolute;top:0;left:0;right:0;height:2px}}
.rl{{font-size:.55rem;font-weight:800;text-transform:uppercase;letter-spacing:1.5px;margin-bottom:6px}}
.rv{{font-size:.8rem;line-height:1.65;white-space:pre-wrap;color:var(--t)}}
.r-cls{{background:{A}08;border:1px solid {A}14}} .r-cls::before{{background:linear-gradient(90deg,{A},transparent)}} .r-cls .rl{{color:{A}}}
.r-pri{{background:rgba(234,179,8,.04);border:1px solid rgba(234,179,8,.10)}} .r-pri::before{{background:linear-gradient(90deg,#eab308,transparent)}} .r-pri .rl{{color:#ca8a04}}
.r-hit{{background:rgba(34,197,94,.04);border:1px solid rgba(34,197,94,.10)}} .r-hit::before{{background:linear-gradient(90deg,#22c55e,transparent)}} .r-hit .rl{{color:#16a34a}}
.r-no{{background:rgba(239,68,68,.04);border:1px solid rgba(239,68,68,.10)}} .r-no::before{{background:linear-gradient(90deg,#ef4444,transparent)}} .r-no .rl{{color:#dc2626}}
.r-fin{{background:linear-gradient(135deg,rgba(34,197,94,.04),{A}06);border:1px solid rgba(34,197,94,.1)}} .r-fin::before{{background:linear-gradient(90deg,#22c55e,{A},transparent)}} .r-fin .rl{{color:#22c55e}}
.r-fallback{{background:rgba(167,139,250,.04);border:1px solid rgba(167,139,250,.18)}} .r-fallback::before{{background:linear-gradient(90deg,#a78bfa,transparent)}} .r-fallback .rl{{color:#a78bfa}}
.sb-logo{{padding:14px 14px 12px;border-bottom:1px solid var(--b)}}
.sb-icon{{width:32px;height:32px;background:linear-gradient(135deg,{A},{A2});border-radius:8px;display:flex;align-items:center;justify-content:center;font-size:16px;box-shadow:0 3px 10px {A}40}}
.sb-name{{font-size:.88rem;font-weight:800;color:var(--t);letter-spacing:-.2px}}
.sb-ver{{font-size:.55rem;color:var(--m);text-transform:uppercase;letter-spacing:1px;font-weight:700}}
.sb-sec{{font-size:.52rem;font-weight:800;color:rgba(255,255,255,.18);text-transform:uppercase;letter-spacing:1.8px;padding:10px 2px 5px}}
.sbc{{display:flex;align-items:center;gap:7px;background:rgba(255,255,255,.02);border:1px solid var(--b);border-radius:7px;padding:7px 9px;margin-bottom:7px}}
.sbd{{width:6px;height:6px;border-radius:50%;flex-shrink:0}}
.sbt{{font-size:.68rem;color:var(--t);font-family:var(--mo);overflow:hidden;text-overflow:ellipsis;white-space:nowrap}}
.sg{{display:grid;grid-template-columns:1fr 1fr;gap:5px;margin-bottom:5px}}
.sc{{background:rgba(255,255,255,.02);border:1px solid var(--b);border-radius:7px;padding:8px;text-align:center}}
.sv{{font-size:1.1rem;font-weight:700;font-family:var(--mo);line-height:1}}
.sl{{font-size:.5rem;color:var(--m);text-transform:uppercase;letter-spacing:.6px;font-weight:700;margin-top:2px}}
.sla{{background:rgba(255,255,255,.015);border:1px solid var(--b);border-radius:7px;overflow:hidden;margin-bottom:7px}}
.sla-r{{display:flex;justify-content:space-between;padding:5px 10px;border-bottom:1px solid rgba(255,255,255,.03);font-size:.7rem}}
.sla-r:last-child{{border-bottom:none}}
.cw{{display:flex;flex-direction:column;gap:8px}}
.cm{{display:flex;gap:8px;animation:ci .25s ease}}
@keyframes ci{{from{{opacity:0;transform:translateY(5px)}}to{{opacity:1;transform:none}}}}
.cm.u{{flex-direction:row-reverse}}
.ca{{width:26px;height:26px;border-radius:6px;flex-shrink:0;display:flex;align-items:center;justify-content:center;font-size:12px;border:1px solid var(--b)}}
.cb{{max-width:84%;padding:8px 11px;border-radius:10px;font-size:.78rem;line-height:1.55;white-space:pre-wrap}}
.cb.u{{background:{A}14;border:1px solid {A}22;color:{A}dd;border-radius:10px 3px 10px 10px}}
.cb.b{{background:var(--c);border:1px solid var(--b);color:var(--t);border-radius:3px 10px 10px 10px}}
.ct{{font-size:.55rem;color:var(--m);margin-top:2px}}
.ct.u{{text-align:right}}
.mg{{display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin-bottom:10px}}
.mt{{background:var(--c);border:1px solid var(--b);border-radius:10px;padding:13px;text-align:center}}
.mt:hover{{border-color:{A}20}}
.mv{{font-size:1.6rem;font-weight:700;font-family:var(--mo);line-height:1}}
.ml{{font-size:.52rem;color:var(--m);text-transform:uppercase;letter-spacing:.7px;font-weight:700;margin-top:4px}}
.hi{{background:var(--c);border:1px solid var(--b);border-radius:8px;padding:9px 12px;margin-bottom:6px;transition:all .18s}}
.hi:hover{{border-color:{A}18;transform:translateX(2px)}}
.ht{{font-family:var(--mo);font-size:.65rem;color:{A};font-weight:600}}
.hiss{{font-size:.77rem;color:var(--t);font-weight:600;margin:2px 0;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.hm{{font-size:.63rem;color:var(--m);display:flex;gap:7px;flex-wrap:wrap}}
hr{{border-color:var(--b)!important}}
</style>
""",unsafe_allow_html=True)

# ── SIDEBAR ──────────────────────────────────────────────────
with st.sidebar:
    st.markdown(f"""
    <div class="sb-logo">
      <div style="display:flex;align-items:center;gap:8px;margin-bottom:6px">
        <div class="sb-icon">🎯</div>
        <div><div class="sb-name">SupportAI</div><div class="sb-ver">Reply Generator · v3.0</div></div>
      </div>
      <div style="font-size:.64rem;color:var(--m);line-height:1.5">LangGraph · RAG · Groq LLaMA 3.1<br>Auto Priority from Past Tickets · Email Fallback</div>
    </div>""",unsafe_allow_html=True)

    st.markdown('<div class="sb-sec">📚 Knowledge Base</div>',unsafe_allow_html=True)
    kb_ok="✅" in st.session_state.kb_status
    kc="#4ade80" if kb_ok else "#fbbf24"
    kt=st.session_state.kb_status[:28]+("…" if len(st.session_state.kb_status)>28 else "")
    st.markdown(f'<div class="sbc"><div class="sbd" style="background:{kc};box-shadow:0 0 5px {kc}90"></div><span class="sbt">{kt}</span></div>',unsafe_allow_html=True)
    upf=st.file_uploader("KB PDF",type=["pdf"],label_visibility="collapsed")
    if upf:
        with st.spinner("Building..."):
            import tempfile as _t2
            with _t2.NamedTemporaryFile(delete=False,suffix=".pdf") as tmp:
                tmp.write(upf.read()); tp=tmp.name
            nv=build_vs(tp); st.session_state.vstore=nv
            st.session_state.graph=build_graph(nv)
            st.session_state.kb_status=f"✅ {upf.name}"; init_excel()
        st.success("✅ KB ready!")

    st.markdown('<div class="sb-sec">📧 Email Config</div>',unsafe_allow_html=True)
    em_ok=bool(SMTP_EMAIL and SMTP_PASS)
    ec="#4ade80" if em_ok else "#fbbf24"
    et=SMTP_EMAIL[:20]+"…" if em_ok and len(SMTP_EMAIL)>20 else (SMTP_EMAIL or "Not configured")
    st.markdown(f'<div class="sbc"><div class="sbd" style="background:{ec};box-shadow:0 0 5px {ec}90"></div><span class="sbt">{et}</span></div>',unsafe_allow_html=True)

    tc=st.session_state.total; rc=st.session_state.resolved
    ec2=st.session_state.escalated; es=st.session_state.emails_sent
    rr=int((rc/tc)*100) if tc>0 else 0
    avg=round(sum(st.session_state.scores)/len(st.session_state.scores),1) if st.session_state.scores else "—"
    st.markdown('<div class="sb-sec">📊 Stats</div>',unsafe_allow_html=True)
    st.markdown(f"""<div class="sg">
      <div class="sc"><div class="sv" style="color:{A}">{tc}</div><div class="sl">Total</div></div>
      <div class="sc"><div class="sv" style="color:#4ade80">{rr}%</div><div class="sl">Resolved</div></div>
      <div class="sc"><div class="sv" style="color:#fb923c">{ec2}</div><div class="sl">Escalated</div></div>
      <div class="sc"><div class="sv" style="color:#a78bfa">{es}</div><div class="sl">Emails</div></div>
      <div class="sc"><div class="sv" style="color:#fbbf24">{avg}</div><div class="sl">Rating</div></div>
      <div class="sc"><div class="sv" style="color:{A2}">50</div><div class="sl">KB</div></div>
    </div>""",unsafe_allow_html=True)

    st.markdown('<div class="sb-sec">⏰ SLA</div>',unsafe_allow_html=True)
    st.markdown(f"""<div class="sla">
      <div class="sla-r"><span>🔴 Critical</span><span style="color:#fca5a5;font-family:var(--mo);font-size:.65rem">2 hrs</span></div>
      <div class="sla-r"><span>🟠 High</span><span style="color:#fdba74;font-family:var(--mo);font-size:.65rem">4 hrs</span></div>
      <div class="sla-r"><span>🟡 Medium</span><span style="color:#fde68a;font-family:var(--mo);font-size:.65rem">24 hrs</span></div>
      <div class="sla-r"><span>🟢 Low</span><span style="color:#86efac;font-family:var(--mo);font-size:.65rem">48 hrs</span></div>
      <div class="sla-r"><span>🤖 Automatic</span><span style="color:{A};font-family:var(--mo);font-size:.65rem">LLM+History</span></div>
    </div>""",unsafe_allow_html=True)

    st.markdown('<div class="sb-sec">💾 Export</div>',unsafe_allow_html=True)
    if Path(EXCEL_PATH).exists():
        with open(EXCEL_PATH,"rb") as f:
            st.download_button("⬇️ Excel Log",f.read(),"support_tickets.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
    if Path(HISTORY_PATH).exists():
        st.download_button("⬇️ JSON History",Path(HISTORY_PATH).read_bytes(),
            "history.json","application/json",use_container_width=True)

# ── MAIN ─────────────────────────────────────────────────────
L,R=st.columns([2.7,1],gap="large")

with L:
    st.markdown(f"""
    <div class="hero">
      <div class="hero-badge"><div class="hero-dot"></div>{IND} · AI Support Live</div>
      <div class="hero-h">Automated <em>Support Reply</em> Generator</div>
      <div class="hero-sub">RAG from past tickets · LLM auto-priority · Email fallback · 8 industries</div>
      <div class="kpi">
        <div class="kp"><div class="kv" style="color:{A}">{tc}</div><div class="kl">Tickets</div></div>
        <div class="kp"><div class="kv" style="color:#4ade80">{rr}%</div><div class="kl">Resolved</div></div>
        <div class="kp"><div class="kv" style="color:#fb923c">{ec2}</div><div class="kl">Escalated</div></div>
        <div class="kp"><div class="kv" style="color:#a78bfa">{es}</div><div class="kl">Emails</div></div>
        <div class="kp"><div class="kv" style="color:#fbbf24">{avg}</div><div class="kl">Rating</div></div>
      </div>
    </div>""",unsafe_allow_html=True)

    T1,T2,T3,T4=st.tabs(["🎫 New Ticket","💬 Chat","📊 Analytics","📋 History"])

    with T1:
        c1,c2,c3=st.columns(3)
        with c1:
            st.markdown(f'<div class="lbl"><div class="dot"></div>{PROF["label"]} Name</div>',unsafe_allow_html=True)
            nm=st.text_input("_n",value=st.session_state.name,placeholder="e.g. Arjun Kumar",label_visibility="collapsed",key="k_name")
            if nm!=st.session_state.name: st.session_state.name=nm
        with c2:
            st.markdown('<div class="lbl"><div class="dot"></div>Email</div>',unsafe_allow_html=True)
            em=st.text_input("_e",value=st.session_state.email,placeholder="email@example.com",label_visibility="collapsed",key="k_email")
            if em!=st.session_state.email: st.session_state.email=em
        with c3:
            st.markdown(f'<div class="lbl"><div class="dot"></div>{PROF["code_label"]}</div>',unsafe_allow_html=True)
            cd=st.text_input("_c",value=st.session_state.code,placeholder="e.g. TN0145",label_visibility="collapsed",key="k_code")
            if cd!=st.session_state.code: st.session_state.code=cd

        st.markdown('<div style="height:4px"></div>',unsafe_allow_html=True)
        p1,p2=st.columns(2)
        with p1:
            st.markdown('<div class="lbl"><div class="dot"></div>Category</div>',unsafe_allow_html=True)
            cat_idx=min(st.session_state.cat_idx,len(CATS)-1)
            cat_sel=st.selectbox("_cat",CATS,index=cat_idx,label_visibility="collapsed",key="k_cat")
            st.session_state.cat_idx=CATS.index(cat_sel)
        with p2:
            st.markdown('<div class="lbl"><div class="dot"></div>Priority</div>',unsafe_allow_html=True)
            pri_sel=st.selectbox("_pri",PRIORITIES,
                index=PRIORITIES.index(st.session_state.priority) if st.session_state.priority in PRIORITIES else 2,
                label_visibility="collapsed",key="k_pri")
            if pri_sel!=st.session_state.priority: st.session_state.priority=pri_sel; reset(); st.rerun()

        if "🤖" in st.session_state.priority:
            hist_count=len(load_history())
            if hist_count>0:
                st.info(f"🤖 **Automatic** — LLM will assign priority based on **{hist_count} past tickets** in history for context-aware triage.")
            else:
                st.info("🤖 **Automatic** — LLM will assign priority using ticket content (no past history yet).")

        tips=QUICK_TIPS.get(cat_sel,[])
        if tips:
            st.markdown('<div class="lbl" style="margin-top:8px"><div class="dot" style="background:#fbbf24"></div>Quick Fill</div>',unsafe_allow_html=True)
            tc_cols=st.columns(len(tips))
            for i,tip in enumerate(tips):
                with tc_cols[i]:
                    if st.button(f"↗ {tip}",key=f"t{i}",use_container_width=True):
                        st.session_state.ticket_input=tip; st.rerun()

        st.markdown('<div class="lbl" style="margin-top:8px"><div class="dot" style="background:#ef4444"></div>Issue Description</div>',unsafe_allow_html=True)
        ti=st.text_area("_t",value=st.session_state.ticket_input,
            placeholder="Describe the issue clearly…\n• What happened?\n• Steps tried?\n• Error shown?",
            height=100,label_visibility="collapsed",key="k_ti")

        st.markdown('<div class="lbl" style="margin-top:8px"><div class="dot" style="background:#94a3b8"></div>Screenshot <span style="opacity:.5;font-size:.55rem;text-transform:none;letter-spacing:0;margin-left:2px">(optional)</span></div>',unsafe_allow_html=True)
        upimg=st.file_uploader("_img",type=["png","jpg","jpeg","webp"],label_visibility="collapsed",key="k_img")
        if upimg:
            ib=upimg.read(); st.session_state.img_b64=base64.b64encode(ib).decode()
            st.session_state.img_name=upimg.name
            st.image(ib,caption=upimg.name,use_column_width=True)
        if st.session_state.img_b64 and not upimg:
            st.markdown(f'<div class="chip" style="color:#fbbf24;border-color:#fbbf2430;background:#fbbf2410;margin-top:4px">📎 {st.session_state.img_name}</div>',unsafe_allow_html=True)

        cr,_=st.columns([1,5])
        with cr:
            if st.button("↺ Reset",key="k_rst"):
                reset(); st.session_state.ticket_input=""; st.session_state.img_b64=None; st.rerun()

        s=st.session_state.current_step
        def nc(n): return "pn pn-d" if n<s else ("pn pn-a" if n==s else "pn pn-i")
        def ni(n): return f'<span class="pi">{"✓" if n<s else n}</span>'
        st.markdown(f"""
        <div class="pipe">
          <div class="pn pn-s">START</div><span class="arr">›</span>
          <div class="{nc(1)}">{ni(1)} Classify</div><span class="arr">›</span>
          <div class="{nc(2)}">{ni(2)} Priority</div><span class="arr">›</span>
          <div class="{nc(3)}">{ni(3)} Retrieve</div><span class="arr">›</span>
          <div class="{nc(4)}">{ni(4)} Generate</div><span class="arr">›</span>
          <div class="{nc(5)}">{ni(5)} Polish</div><span class="arr">›</span>
          <div class="pn pn-e">EMAIL ✉</div>
        </div>""",unsafe_allow_html=True)

        st.markdown('<div class="run-btn">',unsafe_allow_html=True)
        can_run=bool(ti.strip()) and st.session_state.graph is not None
        run=st.button("⚡  Generate Reply  →  Classify · Priority · Retrieve · Generate · Email",
            disabled=not can_run,use_container_width=True,key="k_run")
        st.markdown('</div>',unsafe_allow_html=True)

        if run:
            st.session_state.ticket_input=ti
            st.session_state.current_step=1
            with st.spinner("⚙️ Running pipeline..."):
                fs=st.session_state.graph.invoke({
                    "ticket":ti,"user_name":st.session_state.name or "User",
                    "user_email":st.session_state.email or "N/A",
                    "user_code":st.session_state.code or "N/A",
                    "category":cat_sel,"priority":pri_sel,
                    "auto_priority":"","context":"","kb_matched":False,
                    "raw_response":"","final_response":"","ticket_id":"",
                    "email_fallback":False
                })
            for k in ["category","auto_priority","kb_matched","final_response","ticket_id","email_fallback"]:
                st.session_state[k]=fs.get(k)
            st.session_state.current_step=6
            st.session_state.total+=1
            kb_hit=fs.get("kb_matched",False)
            if kb_hit: st.session_state.resolved+=1
            else: st.session_state.escalated+=1

            should_email = (st.session_state.email and "@" in st.session_state.email)
            email_fallback_triggered = fs.get("email_fallback", False)

            if should_email or email_fallback_triggered:
                target_email = st.session_state.email
                if email_fallback_triggered and not should_email:
                    st.session_state.email_status=(False,"⚠️ No KB match: email fallback triggered but no user email provided. Please add email for escalation notifications.")
                else:
                    subj_prefix="[Support] Escalation Notice" if email_fallback_triggered else "[Support] Response"
                    with st.spinner("📧 Sending email..."):
                        ok,msg=send_email(
                            target_email,st.session_state.name or "User",
                            ti,fs.get("final_response",""),
                            fs.get("ticket_id","TKT"),
                            PROF["greeting"],PROF["sign"],PROF["org"],
                            st.session_state.img_b64,st.session_state.img_name,
                            subject_prefix=subj_prefix
                        )
                    st.session_state.email_status=(ok,msg)
                    if ok: st.session_state.emails_sent+=1

            now=datetime.datetime.now().strftime("%H:%M")
            st.session_state.chat+=[
                {"role":"user","text":ti,"time":now},
                {"role":"bot","text":fs.get("final_response",""),"time":now}
            ]
            save_history({
                "ticket_id":fs.get("ticket_id") or f"TKT-{int(time.time())%9999999:07d}",
                "timestamp":datetime.datetime.now().isoformat(),
                "user_name":st.session_state.name,"user_email":st.session_state.email,
                "user_code":st.session_state.code,"industry":IND,
                "category":fs.get("category",""),"priority":pri_sel,
                "auto_priority":fs.get("auto_priority",""),
                "issue":ti[:200],"status":"Resolved" if kb_hit else "Escalated",
                "email_fallback":email_fallback_triggered,
            })
            st.session_state.history=load_history(); st.rerun()

        if st.session_state.category:
            st.markdown(f'<div class="rc r-cls"><div class="rl">🔵 Classified</div><div class="rv"><b>{html.escape(st.session_state.category)}</b></div></div>',unsafe_allow_html=True)
        if st.session_state.auto_priority:
            was_auto="🤖" in (st.session_state.priority or "")
            hist_c=len(load_history())
            lbl=f"🤖 LLM (based on {hist_c} past tickets) → {st.session_state.auto_priority}" if was_auto else f"Manual: {st.session_state.auto_priority}"
            st.markdown(f'<div class="rc r-pri"><div class="rl">🟡 Priority</div><div class="rv">{html.escape(lbl)}</div></div>',unsafe_allow_html=True)
        if st.session_state.kb_matched is not None:
            if st.session_state.kb_matched:
                st.markdown('<div class="rc r-hit"><div class="rl">🟢 KB Match</div><div class="rv">✅ Similar ticket found in knowledge base — reply generated from past resolutions</div></div>',unsafe_allow_html=True)
            else:
                st.markdown('<div class="rc r-no"><div class="rl">🔴 No KB Match</div><div class="rv">❌ No matching ticket found — escalated to specialist team · Email fallback triggered</div></div>',unsafe_allow_html=True)
        if st.session_state.email_fallback:
            st.markdown('<div class="rc r-fallback"><div class="rl">📧 Email Fallback</div><div class="rv">No KB match detected — escalation email automatically dispatched to user with ticket details and ETA.</div></div>',unsafe_allow_html=True)
        if st.session_state.final_response:
            st.markdown(f'<div class="rc r-fin"><div class="rl">✅ Final Reply</div><div class="rv">{html.escape(st.session_state.final_response)}</div></div>',unsafe_allow_html=True)
            if st.session_state.kb_matched: st.success("✅ Resolved from knowledge base!")
            else: st.warning("📋 Escalated — specialist assigned. Download log from sidebar.")
            if st.session_state.email_status:
                ok,msg=st.session_state.email_status
                (st.success if ok else st.warning)(msg)
            if st.button("📋 Copy Reply",key="k_copy"):
                st.code(st.session_state.final_response,language=None)

            st.markdown('<div class="lbl" style="margin-top:10px;justify-content:center"><div class="dot" style="background:#fbbf24"></div>Rate this reply</div>',unsafe_allow_html=True)
            sc=st.columns(5)
            for i,s in enumerate(["1★","2★","3★","4★","5★"]):
                with sc[i]:
                    if st.button(s,key=f"s{i}",use_container_width=True):
                        st.session_state.scores.append(i+1); st.success(f"Rated {i+1}/5"); st.rerun()

    with T2:
        msgs=st.session_state.chat
        if not msgs:
            st.markdown('<div style="text-align:center;padding:40px 0"><div style="font-size:2rem;margin-bottom:8px">💬</div><div style="font-size:.88rem;font-weight:700;color:var(--t);margin-bottom:4px">No conversation yet</div><div style="font-size:.75rem;color:var(--m)">Generate a reply in New Ticket tab</div></div>',unsafe_allow_html=True)
        else:
            st.markdown('<div class="cw">',unsafe_allow_html=True)
            for m in msgs:
                r=m["role"]; txt=html.escape(m.get("text","")); t=m.get("time","")
                if r=="user":
                    st.markdown(f'<div class="cm u"><div class="ca" style="background:{A}12">👤</div><div><div class="cb u">{txt}</div><div class="ct u">{t}</div></div></div>',unsafe_allow_html=True)
                else:
                    st.markdown(f'<div class="cm"><div class="ca" style="background:rgba(74,222,128,.07)">🤖</div><div><div class="cb b">{txt}</div><div class="ct">{t}</div></div></div>',unsafe_allow_html=True)
            st.markdown('</div>',unsafe_allow_html=True)
            if st.button("🗑️ Clear",key="k_clr"): st.session_state.chat=[]; st.rerun()

    with T3:
        avg2=round(sum(st.session_state.scores)/len(st.session_state.scores),1) if st.session_state.scores else 0
        st.markdown(f"""<div class="mg">
          <div class="mt"><div class="mv" style="color:{A}">{tc}</div><div class="ml">Total</div></div>
          <div class="mt"><div class="mv" style="color:#4ade80">{rc}</div><div class="ml">Resolved</div></div>
          <div class="mt"><div class="mv" style="color:#fb923c">{ec2}</div><div class="ml">Escalated</div></div>
          <div class="mt"><div class="mv" style="color:#fbbf24">{avg2}/5</div><div class="ml">Avg Rating</div></div>
          <div class="mt"><div class="mv" style="color:#a78bfa">{es}</div><div class="ml">Emails</div></div>
          <div class="mt"><div class="mv" style="color:{A2}">50</div><div class="ml">KB Tickets</div></div>
          <div class="mt"><div class="mv" style="color:#f472b6">{len(st.session_state.scores)}</div><div class="ml">Ratings</div></div>
          <div class="mt"><div class="mv" style="color:#34d399">{len(st.session_state.history)}</div><div class="ml">History</div></div>
        </div>""",unsafe_allow_html=True)
        st.markdown(f"""<div class="g">
          <div style="font-size:.65rem;font-weight:700;color:var(--m);margin-bottom:7px">Resolution Rate</div>
          <div style="height:8px;background:var(--b);border-radius:100px;overflow:hidden;margin-bottom:6px">
            <div style="height:100%;width:{rr}%;background:linear-gradient(90deg,#4ade80,{A2});border-radius:100px;transition:width .6s"></div>
          </div>
          <div style="display:flex;justify-content:space-between;font-size:.68rem">
            <span style="color:#4ade80;font-weight:700">✅ {rc} ({rr}%)</span>
            <span style="color:#fb923c;font-weight:700">⚡ {ec2} ({int((ec2/tc)*100) if tc>0 else 0}%)</span>
          </div>
        </div>""",unsafe_allow_html=True)
        if st.session_state.scores:
            st.markdown('<div class="g"><div style="font-size:.65rem;font-weight:700;color:var(--m);margin-bottom:8px">Satisfaction</div>',unsafe_allow_html=True)
            for i in range(5,0,-1):
                cnt=st.session_state.scores.count(i)
                pct=int((cnt/len(st.session_state.scores))*100)
                st.markdown(f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:5px"><span style="font-size:.68rem;color:var(--m);width:22px;text-align:right;font-weight:700">{i}★</span><div style="flex:1;height:6px;background:var(--b);border-radius:100px;overflow:hidden"><div style="height:100%;width:{pct}%;background:linear-gradient(90deg,#fbbf24,#f59e0b);border-radius:100px"></div></div><span style="font-size:.62rem;color:var(--m);width:22px;font-family:var(--mo)">{cnt}×</span></div>',unsafe_allow_html=True)
            st.markdown('</div>',unsafe_allow_html=True)

    with T4:
        hist=st.session_state.history
        if not hist:
            st.markdown('<div style="text-align:center;padding:40px 0"><div style="font-size:2rem;margin-bottom:8px">📋</div><div style="font-size:.88rem;font-weight:700;color:var(--t);margin-bottom:4px">No history yet</div><div style="font-size:.75rem;color:var(--m)">Submitted tickets appear here</div></div>',unsafe_allow_html=True)
        else:
            st.markdown('<div class="lbl"><div class="dot"></div>Search</div>',unsafe_allow_html=True)
            sq=st.text_input("_sq",placeholder="Search tickets, names, categories…",label_visibility="collapsed",key="k_sq")
            fl=hist
            if sq:
                q=sq.lower()
                fl=[h for h in hist if any(q in str(h.get(k,"")).lower() for k in ["ticket_id","issue","user_name","category","industry"])]
            st.markdown(f'<div style="font-size:.62rem;color:var(--m);margin-bottom:8px">Showing <b style="color:var(--t)">{len(fl)}</b> of {len(hist)}</div>',unsafe_allow_html=True)
            for h in fl[:30]:
                tid=h.get("ticket_id","—"); issue=h.get("issue","")[:75]
                cat=h.get("category",""); pri=h.get("priority",""); ap=h.get("auto_priority","")
                st2=h.get("status",""); uname=h.get("user_name",""); ind2=h.get("industry","")
                ef=h.get("email_fallback",False)
                ts_raw=h.get("timestamp","")
                try: ts=datetime.datetime.fromisoformat(ts_raw).strftime("%d %b %y %H:%M")
                except: ts=ts_raw[:16]
                sc2="#4ade80" if st2=="Resolved" else "#fb923c"
                pc=PCOLORS.get(pri,"#94a3b8")
                fb_badge=f'<span class="chip" style="color:#a78bfa;border-color:#a78bfa30;background:#a78bfa10">📧 Fallback</span>' if ef else ""
                st.markdown(f"""<div class="hi">
                  <div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:3px">
                    <span class="ht">{html.escape(tid)}</span>
                    <div style="display:flex;gap:4px;flex-wrap:wrap">
                      <span class="chip" style="color:{pc};border-color:{pc}30;background:{pc}10">{html.escape(pri)}</span>
                      {f'<span class="chip" style="color:{A};border-color:{A}25;background:{A}0e">🤖 {html.escape(ap)}</span>' if ap and ap!=pri else ""}
                      <span class="chip" style="color:{sc2};border-color:{sc2}30;background:{sc2}10">{html.escape(st2)}</span>
                      {fb_badge}
                    </div>
                  </div>
                  <div class="hiss">{html.escape(issue+"…" if len(h.get("issue",""))>75 else issue)}</div>
                  <div class="hm">
                    <span>{html.escape(ind2)}</span>
                    <span>📂 {html.escape(cat)}</span>
                    <span>👤 {html.escape(uname)}</span>
                    <span>🕐 {ts}</span>
                  </div>
                </div>""",unsafe_allow_html=True)

with R:
    st.markdown('<div style="height:4px"></div>',unsafe_allow_html=True)
    st.markdown(f"""<div style="display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-bottom:8px">
      <div style="background:{A}08;border:1px solid {A}14;border-radius:10px;padding:10px;text-align:center">
        <div style="font-size:1.25rem;font-weight:700;color:{A};font-family:var(--mo)">{tc}</div>
        <div style="font-size:.5rem;color:var(--m);text-transform:uppercase;letter-spacing:.7px;font-weight:700;margin-top:2px">Total</div>
      </div>
      <div style="background:rgba(74,222,128,.05);border:1px solid rgba(74,222,128,.12);border-radius:10px;padding:10px;text-align:center">
        <div style="font-size:1.25rem;font-weight:700;color:#4ade80;font-family:var(--mo)">{rr}%</div>
        <div style="font-size:.5rem;color:var(--m);text-transform:uppercase;letter-spacing:.7px;font-weight:700;margin-top:2px">Resolved</div>
      </div>
      <div style="background:rgba(251,146,60,.05);border:1px solid rgba(251,146,60,.12);border-radius:10px;padding:10px;text-align:center">
        <div style="font-size:1.25rem;font-weight:700;color:#fb923c;font-family:var(--mo)">{ec2}</div>
        <div style="font-size:.5rem;color:var(--m);text-transform:uppercase;letter-spacing:.7px;font-weight:700;margin-top:2px">Escalated</div>
      </div>
      <div style="background:rgba(167,139,250,.05);border:1px solid rgba(167,139,250,.12);border-radius:10px;padding:10px;text-align:center">
        <div style="font-size:1.25rem;font-weight:700;color:#a78bfa;font-family:var(--mo)">{es}</div>
        <div style="font-size:.5rem;color:var(--m);text-transform:uppercase;letter-spacing:.7px;font-weight:700;margin-top:2px">Emails</div>
      </div>
    </div>""",unsafe_allow_html=True)

    if st.session_state.ticket_input:
        ap_d=st.session_state.auto_priority or "—"
        rows="".join(f'<div style="display:flex;justify-content:space-between;padding:3px 0;border-bottom:1px solid rgba(255,255,255,.03);font-size:.68rem"><span style="color:var(--m)">{k}</span><span style="color:var(--t);font-weight:600;font-family:var(--mo);font-size:.64rem">{v}</span></div>'
            for k,v in [
                ("Industry",IND.split()[-1]),
                ("Name",(st.session_state.name or "—")[:11]),
                ("Code",st.session_state.code or "—"),
                ("Category",(cat_sel or "—")[:13]),
                ("Priority",pri_sel),
                ("Auto-P",ap_d),
                ("Step",f"{st.session_state.current_step}/5"),
            ])
        st.markdown(f'<div class="g" style="margin-bottom:8px"><div style="font-size:.52rem;font-weight:800;text-transform:uppercase;letter-spacing:1.8px;color:rgba(255,255,255,.18);margin-bottom:7px">🎫 Active Ticket</div>{rows}</div>',unsafe_allow_html=True)

    st.markdown('<div class="g" style="margin-bottom:8px">',unsafe_allow_html=True)
    st.markdown('<div style="font-size:.52rem;font-weight:800;text-transform:uppercase;letter-spacing:1.8px;color:rgba(255,255,255,.18);margin-bottom:8px">🤖 Pipeline Steps</div>',unsafe_allow_html=True)
    for n,col,nm,desc in [
        ("1",A,"Classify","Route to category"),
        ("2","#fbbf24","Priority","LLM + past tickets"),
        ("3",A2,"Retrieve","RAG search"),
        ("4","#4ade80","Generate","Draft / Escalate"),
        ("5","#a78bfa","Polish","Final check"),
    ]:
        done=int(n)<st.session_state.current_step
        active=int(n)==(st.session_state.current_step or 0)
        bd=f"border-color:{col}" if active else "border-color:var(--b)"
        bg=f"background:color-mix(in srgb,{col} 7%,transparent)" if (done or active) else "background:transparent"
        st.markdown(f'<div style="display:flex;align-items:center;gap:7px;padding:5px 7px;border-radius:7px;margin-bottom:3px;border:1px solid;{bd};{bg};transition:all .18s"><div style="width:18px;height:18px;background:color-mix(in srgb,{col} 14%,transparent);border:1.5px solid color-mix(in srgb,{col} 35%,transparent);border-radius:4px;display:flex;align-items:center;justify-content:center;font-size:.52rem;font-weight:900;color:{col};flex-shrink:0">{"✓" if done else n}</div><div><div style="font-size:.68rem;font-weight:700;color:var(--t)">{nm}</div><div style="font-size:.57rem;color:var(--m)">{desc}</div></div></div>',unsafe_allow_html=True)
    st.markdown('</div>',unsafe_allow_html=True)

    st.markdown('<div class="g">',unsafe_allow_html=True)
    st.markdown('<div style="font-size:.52rem;font-weight:800;text-transform:uppercase;letter-spacing:1.8px;color:rgba(255,255,255,.18);margin-bottom:8px">✨ Features</div>',unsafe_allow_html=True)
    for icon,feat in [
        ("🤖","LLM Auto Priority + History"),("📚","RAG Knowledge Base"),
        ("📧","Email Fallback on Escalate"),("💬","Chat History"),
        ("📊","Live Analytics"),("📋","Ticket Log"),
        ("🏢","8 Industries"),("📎","Screenshot Support"),
    ]:
        st.markdown(f'<div style="display:flex;align-items:center;gap:6px;margin-bottom:4px;font-size:.68rem;color:var(--m)"><span style="font-size:.8rem">{icon}</span>{feat}</div>',unsafe_allow_html=True)
    st.markdown('</div>',unsafe_allow_html=True)
